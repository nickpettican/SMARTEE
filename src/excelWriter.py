#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ___        SMARTEE V 0.0.1 by nickpettican            ___
# ___        Social Media Automated Research            ___
# ___        Tool for Evaluating Engagement             ___

# ___        Copyright 2017 Nicolas Pettican			___

# ___    This software is licensed under the Apache 2	___
# ___    license. You may not use this file except in 	___
# ___    compliance with the License.					___
# ___    You may obtain a copy of the License at 		___

# ___    http://www.apache.org/licenses/LICENSE-2.0		___

# ___    Unless required by applicable law or agreed	___
# ___    to in writing, software distributed under		___
# ___    the License is distributed on an "AS IS"		___
# ___    BASIS, WITHOUT WARRANTIES OR CONDITIONS OF 	___
# ___    ANY KIND, either express or implied. See the 	___
# ___    License for the specific language governing	___
# ___    permissions and limitations under the License.	___

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart.axis import DateAxis
from openpyxl.chart import LineChart, Reference
from os import path, mkdir
from collections import Counter
from datetime import date as d
import arrow

class outputExcel:

	def __init__(self, mode, console):
		# initialise Excel writer with the mode

		self.mode = mode
		self.console = console
		self.check_path()

	def check_path(self):
		# create folder if not present

		if not path.isdir('results'):
			mkdir('results')

		self.day_path = 'results/' + str(arrow.now().format('DD_MM_YYYY/'))

		if not path.isdir(self.day_path):
			mkdir(self.day_path)

		self.out_path = 'results/' + str(arrow.now().format('DD_MM_YYYY/')) + 'Instagram/'

		if not path.isdir(self.out_path):
			mkdir(self.out_path)

	def prepareData(self, data):
		# organises the data into an easily writable format
		# matrix of posts as the rows, columns being the data fields in those posts
		# organised by username and popularity

		header = [['Username', 'Followers', 'Following', 'Media', 'Date Published', 'Time Published', 'Timestamp', 'Popularity Score', 'Likes', 'Comments', 'Caption', 'URL', 'Image URL', 'Hashtags']]
		final = {key: header for key in data['hashtags'].keys()}
		order = ['date', 'time', 'timestamp', 'popularity', 'likes_count', 'comments_count', 'caption', 'url', 'display_src', 'hashtags']

		for hashtag, user_data_list in data['hashtags'].items():
			for user_data in user_data_list:
				rows = [[user_data.keys()[0], user_data.values()[0]['followers'], user_data.values()[0]['following'], user_data.values()[0]['media']['count']] +
						[post_data[i] if type(post_data[i]) != list else ', '.join(post_data[i]) for i in order] for post_data in user_data.values()[0]['media']['nodes'] if post_data]

				sorted_rows = sorted(rows, key = lambda x: -x[7])
				final[hashtag].extend(sorted_rows)

		return final

	def dateAnalytics(self, data):
		# returns the number of posts per day

		final = {hashtag: [] for hashtag in data.keys()}
		for hashtag, matrix in data.items():
			matrix = matrix[1:]
			table = [['Date', hashtag.title()]]
			matrix_sorted = sorted(matrix, key = lambda x: x[6])
			# count the posts for each day
			date = matrix_sorted[0][4]
			date_counter = 0
			for i, row in enumerate(matrix_sorted, 1):
				if not row[4] == date or i == len(matrix_sorted):
					dateSplit = [int(i) for i in date.split('-')]
					table.append([d(dateSplit[0], dateSplit[1], dateSplit[2]), date_counter])
					date = row[4]
					date_counter = 1
					continue
				date_counter += 1
			final[hashtag] = table

		return final

	def userAnalytics(self, data):
		# returns users data

		final = {hashtag: [] for hashtag in data.keys()}

		def add_if_int(variable, to_add):
			# in case 'to_add' is not int

			try:
				return int(variable) + int(to_add)

			except TypeError:
				return variable

		for hashtag, matrix in data.items():
			matrix = matrix[1:]
			table = [['User', 'Media', '%s Media' %(hashtag), 'Total Likes for %s' %(hashtag), 'Total Comments for %s' %(hashtag), 
						'Followers', 'Following', 'Avg Popularity', 'Top Hashtags Used']]
			user = matrix[0][0]
			# initialise counters
			media_counter = 0
			total_likes = 0
			total_comments = 0
			popularity = 0
			hashtags_used = []
			# organise data
			for i, post in enumerate(matrix, 1):
				# first round skips this
				# if user changes, append previous user data
				if not post[0] == user:
					table.append([user, old_post[3], media_counter, total_likes, total_comments, old_post[1], old_post[2], popularity/media_counter, 
								', '.join([n[0] for n in Counter(hashtags_used).most_common(10)])])
					# if it is the last user append data
					if i == len(matrix):
						table.append([post[0], post[3], media_counter, total_likes, total_comments, post[1], post[2], popularity/media_counter, 
									', '.join([n[0] for n in Counter(hashtags_used).most_common(10)])])
					# update new user data and save current user data to append later
					user = post[0]
					media_counter = 1
					total_likes = post[8]
					total_comments = post[9]
					popularity = post[7]
					hashtags_used = [n.strip() for n in post[-1].split(',')]
					old_post = post
					continue
				# user is the same, add to counters
				media_counter += 1
				total_likes = add_if_int(total_likes, post[8])
				total_comments = add_if_int(total_comments, post[9])
				popularity = add_if_int(popularity, post[7])
				hashtags_used.extend([n.strip() for n in post[-1].split(',')])
				# still need to keep the data in case next user is new
				old_post = post
				# if it is the same user but the last row
				if i == len(matrix):
					table.append([post[0], post[3], media_counter, total_likes, total_comments, post[1], post[2], popularity/media_counter, 
								', '.join([n[0] for n in Counter(hashtags_used).most_common(10)])])
			final[hashtag] = table

		return final

	def write(self, data):
		# writes the instaCrawl object to xlsx

		self.console.log('Writing to Excel... \,')
		hashtags = data['hashtags'].keys()
		final = self.prepareData(data)
		by_date = self.dateAnalytics(final)
		by_user = self.userAnalytics(final)

		fileName = self.out_path + arrow.now().format('YYYY_MM_DD-HH-mm') + '.xlsx'
		wb = Workbook()
		sheetFirst = wb.active
		chart_position = 2

		# write the graphs to the first sheet
		for hashtag in hashtags:
			try:
				print '# ',
				sheet = wb.create_sheet(title = '#' + hashtag + ' users')
				for row in final[hashtag]:
					sheet.append(row)
				userAnalyticsSheet = wb.create_sheet(title = '#' + hashtag + ' user analytics')
				for row in by_user[hashtag]:
					userAnalyticsSheet.append(row)
				dateAnalyticsSheet = wb.create_sheet(title = '#' + hashtag + ' posts by date')
				for row in by_date[hashtag]:
					dateAnalyticsSheet.append(row)

				# initialise the chart
				chart = LineChart()
				chart.title = '#' + hashtag
				chart.style = 12
				chart.y_axis.title = 'Number of posts'
				chart.y_axis.crossAx = 500
				chart.x_axis = DateAxis(crossAx=100)
				chart.x_axis.number_format = 'd-mmm'
				chart.x_axis.majorTimeUnit = 'days'
				chart.x_axis.title = 'Date'
				# add the data
				chartData = Reference(dateAnalyticsSheet, min_col=2, min_row=1, max_col=2, max_row=len(by_date[hashtag]))
				dates = Reference(dateAnalyticsSheet, min_col=1, min_row=2, max_row=len(by_date[hashtag]))
				chart.add_data(chartData, titles_from_data=True)
				chart.set_categories(dates)
				# style the chart
				line1 = chart.series[0]
				line1.smooth = True
				sheetFirst.add_chart(chart, 'A' + str(chart_position))
				# increase chart position for next one
				chart_position += 15

			except Exception as e:
				self.console.log('Error writing Excel file: %s' %(e))

		wb.save(fileName)
		self.console.log('done.')
